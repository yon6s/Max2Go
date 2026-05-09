from openpyxl import load_workbook
import json
path='/Users/wang/Downloads/Max2Go.xlsx'
wb=load_workbook(path,data_only=True)
ws=wb['租决']
rooms=[]
for row in range(2,170):
    building=ws[f'A{row}'].value
    room=ws[f'B{row}'].value
    area=ws[f'C{row}'].value
    price=ws[f'D{row}'].value
    status=ws[f'E{row}'].value
    if building and room and isinstance(area,(int,float)) and isinstance(price,(int,float)):
        rooms.append({
            'building': str(building),
            'room': str(room),
            'area': round(float(area), 4),
            'price': round(float(price), 6),
            'status': str(status or ''),
        })
pairs=[]
for rate_col, year_col in [('Q','R'),('T','U'),('W','X'),('Z','AA'),('AC','AD')]:
    for row in range(3,54):
        r=ws[f'{rate_col}{row}'].value
        y=ws[f'{year_col}{row}'].value
        if isinstance(r,(int,float)) and isinstance(y,(int,float)):
            pairs.append((round(float(r),6), round(float(y),4)))
seen={}
for r,y in pairs:
    seen.setdefault(r,y)
j_table=[{'rate':r,'years':y} for r,y in sorted(seen.items(), key=lambda x:x[0], reverse=True)]
out='window.MAX2GO_DATA = '+json.dumps({'rooms':rooms,'jTable':j_table},ensure_ascii=False,separators=(',',':'))+';\n'
open('public/assets/max2go-data.js','w',encoding='utf-8').write(out)
print(len(rooms), len(j_table), rooms[0], rooms[-1])

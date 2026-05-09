from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json, re
path='/Users/wang/Downloads/Max2Go.xlsx'
wb = load_workbook(path, data_only=False)
print('SHEETS', wb.sheetnames)
terms=re.compile(r'破底|回正|租决|合同|免租|物业|车位|涨幅|分户|精装|成本|总价|面积|价格|J|底价')
for ws in wb.worksheets:
    print('\n---', ws.title, 'max', ws.max_row, ws.max_column, '---')
    # Print non-empty cells in used range, compact
    count=0
    for row in ws.iter_rows():
        vals=[]
        has=False
        for cell in row:
            v=cell.value
            if v is not None:
                has=True
                val=str(v)
                if len(val)>80: val=val[:77]+'...'
                vals.append(f'{cell.coordinate}={val}')
        if has:
            print(' | '.join(vals))
            count += 1
        if count>120:
            print('... truncated rows')
            break
    print('FORMULAS:')
    fcount=0
    for row in ws.iter_rows():
        for cell in row:
            v=cell.value
            if isinstance(v,str) and v.startswith('='):
                print(f'{cell.coordinate}: {v}')
                fcount += 1
                if fcount>200:
                    print('... truncated formulas')
                    break
        if fcount>200: break
    print('TERM MATCHES:')
    for row in ws.iter_rows():
        for cell in row:
            v=cell.value
            if v is not None and terms.search(str(v)):
                print(f'{cell.coordinate}: {v}')

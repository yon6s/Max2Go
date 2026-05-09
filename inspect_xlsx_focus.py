from openpyxl import load_workbook
path='/Users/wang/Downloads/Max2Go.xlsx'
wb_formula = load_workbook(path, data_only=False)
wb_values = load_workbook(path, data_only=True)
wsf = wb_formula['租决']
wsv = wb_values['租决']
for rng in ['G1:J22','A130:E170','Q1:AD60','Q1:AD170']:
    print('\n--- RANGE', rng, 'FORMULAS/VALUE ---')
    for row in wsf[rng]:
        parts=[]
        for c in row:
            v=c.value
            if v is not None:
                parts.append(f'{c.coordinate}={v}')
        if parts: print(' | '.join(parts))
    print('\n--- RANGE', rng, 'CALC VALUES ---')
    for row in wsv[rng]:
        parts=[]
        for c in row:
            v=c.value
            if v is not None:
                parts.append(f'{c.coordinate}={v}')
        if parts: print(' | '.join(parts))

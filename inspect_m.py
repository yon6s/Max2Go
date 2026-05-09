from openpyxl import load_workbook
path='/Users/wang/Downloads/Max2Go.xlsx'
for data_only in [False, True]:
 print('DATA_ONLY', data_only)
 wb=load_workbook(path,data_only=data_only)
 ws=wb['租决']
 for row in ws['G17:M45']:
  parts=[]
  for c in row:
   if c.value is not None:
    parts.append(f'{c.coordinate}={c.value}')
  if parts: print(' | '.join(parts))

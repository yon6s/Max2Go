from openpyxl import load_workbook
import json
wb=load_workbook('/Users/wang/Downloads/Max2Go.xlsx',data_only=True)
ws=wb['租决']
pairs=[]
for rate_col, year_col in [('Q','R'),('T','U'),('W','X'),('Z','AA'),('AC','AD')]:
    for row in range(3,54):
        r=ws[f'{rate_col}{row}'].value
        y=ws[f'{year_col}{row}'].value
        if isinstance(r,(int,float)) and isinstance(y,(int,float)):
            pairs.append((float(r),float(y)))
# dedupe by rate, keep first then sort desc rate
seen={}
for r,y in pairs:
    seen.setdefault(round(r,6), y)
items=sorted([(r,y) for r,y in seen.items()], key=lambda x:x[0], reverse=True)
print(json.dumps(items,ensure_ascii=False))
print(len(items), items[0], items[-1])
